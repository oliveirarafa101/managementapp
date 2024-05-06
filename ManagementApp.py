from kivy.app import App
from kivy.uix.floatlayout import FloatLayout
from kivy.uix.image import Image
from kivy.uix.textinput import TextInput
from kivy.uix.popup import Popup
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.dropdown import DropDown
from kivy.metrics import dp
from kivy.uix.gridlayout import GridLayout
from clientes import clientes,clientes_faturar
import openpyxl
import os.path

class Root(App):
    def build(self):
        layout = FloatLayout()
        
        # Adiciona a imagem de um caminhão como plano de fundo
        truck_image = Image(source='truck.jpeg', allow_stretch=True, keep_ratio=False)
        layout.add_widget(truck_image)
        
        # Adiciona campos de entrada na frente da imagem, centralizados horizontalmente
        self.inputs = {}
        y_pos = 0.9  # Define a posição vertical inicial das caixas de entrada
        for idx, header in enumerate(["CLIENTE A FATURAR", "CLIENTE", "DATA", "CONTENTOR", "VIATURA", "GUIA TRANSPORTE"]):
            if header == "CLIENTE A FATURAR":
                # Cria um input label para o cliente a faturar com dropdown
                input_label = DropDownTextInput(hint_text=header, size_hint=(0.5, 0.05), pos_hint={'center_x': 0.5, 'top': y_pos}, clients=clientes_faturar)
                layout.add_widget(input_label)
                self.inputs[header] = input_label
            elif header == "CLIENTE": 
                 # Cria um input label para o cliente com dropdown
                input_label = DropDownTextInput(hint_text=header, size_hint=(0.5, 0.05), pos_hint={'center_x': 0.5, 'top': y_pos}, clients=clientes)
                layout.add_widget(input_label)
                self.inputs[header] = input_label
            else:
                input_field = TextInput(hint_text=header, size_hint=(0.5, 0.05), pos_hint={'center_x': 0.5, 'top': y_pos}, background_color=(0.8, 0.8, 0.8, 1))
                layout.add_widget(input_field)
                self.inputs[header] = input_field
            y_pos -= 0.09  # Atualiza a posição vertical para o próximo campo
        
        # Adiciona botão de salvar
        save_button = Button(text="Save to Excel", size_hint=(0.3, 0.1), pos_hint={'center_x': 0.3, 'y': 0.05})
        save_button.bind(on_press=self.save_to_excel)
        layout.add_widget(save_button)

        # Adiciona botão para mostrar os dados já inseridos
        show_data_button = Button(text="Show Inserted Data", size_hint=(0.3, 0.1), pos_hint={'center_x': 0.7, 'y': 0.05})
        show_data_button.bind(on_press=self.show_inserted_data)
        layout.add_widget(show_data_button)
        
        return layout

    def save_to_excel(self, instance):
        file_path = "dados_camioes.xlsx"
        headers = ["CLIENTE A FATURAR", "CLIENTE", "DATA", "CONTENTOR", "VIATURA", "GUIA TRANSPORTE"]  # Movido para fora do bloco if-else
        
        # Verifica se o arquivo já existe
        if not os.path.isfile(file_path):
            # Se não existir, cria um novo arquivo Excel com os cabeçalhos
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(headers)
        else:
            # Se existir, carrega o arquivo Excel existente
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

        # Obtém os dados dos campos de entrada
        data = [self.inputs[header].text for header in headers]

        # Salva o arquivo Excel
        sheet.append(data)
        workbook.save(file_path)

        # Feedback ao usuário
        success_popup = Popup(title='Success', content=Button(text='Data saved successfully!'), size_hint=(None, None), size=(400, 400))
        success_popup.open()

    def show_inserted_data(self, instance):
        file_path = "dados_camioes.xlsx"
        
        # Verifica se o arquivo existe antes de tentar carregá-lo
        if os.path.isfile(file_path):
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            # Mostra os dados já inseridos em uma tabela organizada
            data_grid = GridLayout(cols=len(sheet[1]) + 1, spacing=5, size_hint_y=None, pos_hint={'center_x': 0.5, 'y': 0.2})
            data_grid.bind(minimum_height=data_grid.setter('height'))

            # Adiciona cabeçalhos
            for cell in sheet[1]:
                label = Label(text=str(cell.value), size_hint=(None, None), size=(dp(100), dp(40)), color=(1, 1, 1, 1))
                data_grid.add_widget(label)
            
            # Adiciona um cabeçalho para os botões de edição
            edit_header = Label(text='Edit', size_hint=(None, None), size=(dp(100), dp(40)), color=(1, 1, 1, 1))
            data_grid.add_widget(edit_header)

            # Adiciona dados
            for row in sheet.iter_rows(min_row=2, values_only=True):
                for item in row:
                    label = Label(text=str(item), size_hint=(None, None), size=(dp(100), dp(40)), color=(1, 1, 1, 1))
                    data_grid.add_widget(label)
                
                # Adiciona botão de edição para cada linha
                edit_button = Button(text='Edit', size_hint=(None, None), size=(dp(100), dp(40)))
                edit_button.bind(on_release=lambda btn, row=row: self.edit_data(sheet, row))  # Correção aqui
                data_grid.add_widget(edit_button)

            data_popup = Popup(title='', content=data_grid, size_hint=(None, None), size=(dp(800), dp(400)))
            data_popup.open()
        else:
            # Se o arquivo não existir, exibe uma mensagem de erro
            error_popup = Popup(title='Error', content=Label(text='No data available!'), size_hint=(None, None), size=(400, 200))
            error_popup.open()

    def edit_data(self, sheet, row):  # Correção aqui
        # Cria uma caixa de diálogo popup para editar os dados
        edit_popup = Popup(title='Edit Data', size_hint=(None, None), size=(400, 400))

        # Layout para organizar os campos de edição
        layout = GridLayout(cols=2, spacing=10, padding=10)

        # Adiciona os campos de entrada para cada item na linha
        for item in row:
            input_field = TextInput(text=str(item), size_hint=(None, None), height=40)
            layout.add_widget(Label(text=str(sheet[1][row.index(item)].value) + ':', size_hint=(None, None), height=40))
            layout.add_widget(input_field)

        # Botão para salvar as alterações
        save_button = Button(text='Save Changes', size_hint=(None, None), size=(150, 40))
        layout.add_widget(save_button)

        edit_popup.content = layout
        edit_popup.open()

class DropDownTextInput(TextInput):
    def __init__(self, clients, **kwargs):
        super().__init__(**kwargs)
        self.dropdown = DropDown()  # Cria uma instância de DropDown

        # Lista de clientes
        self.clients = clients

        # Adiciona cada cliente como um botão no DropDown
        for client in self.clients:
            btn = Button(text=client, size_hint_y=None, height=44)
            btn.bind(on_release=lambda btn: self.update_client_field(btn.text))  # Atualiza o campo de cliente ao clicar em um botão
            self.dropdown.add_widget(btn)
        
    def update_client_field(self, text):
        # Atualiza o texto do input label do cliente
        self.text = text
        self.dropdown.dismiss()

    def on_focus(self, instance, value):
        if value:
            # Abre o DropDown abaixo do input label do cliente quando ganha foco
            self.dropdown.open(instance)
        else:
            # Dispensa o DropDown quando o foco é perdido
            self.dropdown.dismiss()

if __name__ == "__main__":
    Root().run()
